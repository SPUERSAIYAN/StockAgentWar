from __future__ import annotations

import math
import re
from functools import lru_cache
from pathlib import Path
from typing import Any
from zipfile import ZipFile
from xml.etree import ElementTree


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONCEPT_BOARD_PATH = PROJECT_ROOT / "astockdate" / "全部A股20264.xlsx"
CONCEPT_BOARD_SOURCE_LABEL = "candidate_discovery.local_concept_board"
CONCEPT_BOARD_SHEET_NAME = "Sheet1"


def discover_local_concept_board_candidates(
    requested_sectors: list[str],
    discovery_config: dict[str, Any],
) -> dict[str, Any]:
    path = resolve_concept_board_path(discovery_config)
    max_candidates = max(int(discovery_config.get("max_candidates", 8)), 1)
    filters = dict(discovery_config.get("a_share_filters", {}) or {})
    base = {
        "method": "local_excel_concept_board",
        "requested_sectors": requested_sectors,
        "matched_sectors": [],
        "matched_count": 0,
        "filtered_out_count": 0,
        "source_labels": [CONCEPT_BOARD_SOURCE_LABEL],
        "sources": {},
        "errors": {},
        "candidates": [],
    }

    try:
        rows = load_concept_board_rows(path)
        matches, matched_sectors = match_concept_board_rows(rows, requested_sectors)
    except Exception as exc:
        return {
            **base,
            "errors": {CONCEPT_BOARD_SOURCE_LABEL: exc},
        }

    source_summary = {
        "path": str(path),
        "sheet_name": CONCEPT_BOARD_SHEET_NAME,
        "row_count": len(rows),
        "requested_sectors": requested_sectors,
        "matched_sectors": matched_sectors,
        "matched_count": len(matches),
    }

    if not matches:
        return {
            **base,
            "matched_sectors": matched_sectors,
            "sources": {CONCEPT_BOARD_SOURCE_LABEL: source_summary},
            "errors": {
                CONCEPT_BOARD_SOURCE_LABEL: LookupError(
                    f"本地概念板块表未匹配到指定板块：{', '.join(requested_sectors)}"
                )
            },
        }

    candidates: list[dict[str, Any]] = []
    filtered_out_count = 0
    for match in matches:
        row = dict(match["row"])
        if not passes_static_filters(row, filters):
            filtered_out_count += 1
            continue
        score = score_static_candidate(row)
        matched_concepts = list(match["matched_concepts"])
        requested = list(match["requested_sectors"])
        sector = matched_concepts[0] if matched_concepts else requested[0]
        candidates.append(
            {
                "symbol": row["symbol"],
                "name": row["name"],
                "market": "CN",
                "reason": build_static_candidate_reason(
                    requested=requested,
                    matched_concepts=matched_concepts,
                    score=score,
                ),
                "score": score,
                "metadata": {
                    "sector": sector,
                    "matched_concepts": matched_concepts,
                    "requested_sectors": requested,
                    "industry": row.get("industry", ""),
                    "concepts": row.get("concepts", ""),
                    "revenue_growth_yoy": row.get("revenue_growth_yoy"),
                    "net_profit_growth_yoy": row.get("net_profit_growth_yoy"),
                    "net_profit_ytd_yi": row.get("net_profit_ytd_yi"),
                    "roe": row.get("roe"),
                    "pe": row.get("pe"),
                    "forecast_pe_2026": row.get("forecast_pe_2026"),
                    "pe_median_since_2023": row.get("pe_median_since_2023"),
                    "forecast_net_profit_growth_yoy": row.get("forecast_net_profit_growth_yoy"),
                    "total_market_cap_cny_100m": row.get("total_market_cap_cny_100m"),
                    "average_trading_price_2025": row.get("average_trading_price_2025"),
                    "source": "local_excel_concept_board",
                },
            }
        )

    candidates = sorted(
        candidates,
        key=lambda item: (
            item.get("score") or 0,
            item.get("metadata", {}).get("total_market_cap_cny_100m") or 0,
        ),
        reverse=True,
    )
    result = {
        **base,
        "matched_sectors": matched_sectors,
        "matched_count": len(matches),
        "filtered_out_count": filtered_out_count,
        "sources": {
            CONCEPT_BOARD_SOURCE_LABEL: {
                **source_summary,
                "filtered_out_count": filtered_out_count,
                "candidate_count": min(len(candidates), max_candidates),
            }
        },
        "candidates": candidates[:max_candidates],
    }
    if not result["candidates"]:
        result["errors"] = {
            CONCEPT_BOARD_SOURCE_LABEL: LookupError(
                "本地概念板块表有匹配股票，但按候选过滤条件过滤后无可用候选。"
            )
        }
    return result


def resolve_concept_board_path(discovery_config: dict[str, Any]) -> Path:
    raw_path = (
        discovery_config.get("local_concept_board_path")
        or discovery_config.get("a_share_concept_board_path")
        or discovery_config.get("concept_board_path")
        or DEFAULT_CONCEPT_BOARD_PATH
    )
    path = Path(str(raw_path))
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


def load_concept_board_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"本地 A 股概念板块文件不存在：{path}")
    stat = path.stat()
    return [dict(row) for row in _load_concept_board_rows_cached(str(path), stat.st_mtime_ns)]


@lru_cache(maxsize=4)
def _load_concept_board_rows_cached(path_text: str, mtime_ns: int) -> tuple[dict[str, Any], ...]:
    del mtime_ns
    try:
        import openpyxl
    except ImportError as exc:
        del exc
        raw_rows = read_xlsx_sheet_values(Path(path_text), CONCEPT_BOARD_SHEET_NAME)
    else:
        workbook = openpyxl.load_workbook(path_text, read_only=True, data_only=True)
        if CONCEPT_BOARD_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"本地 A 股概念板块表缺少工作表：{CONCEPT_BOARD_SHEET_NAME}")
        sheet = workbook[CONCEPT_BOARD_SHEET_NAME]
        raw_rows = list(sheet.iter_rows(values_only=True))

    rows = iter(raw_rows)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError("本地 A 股概念板块表为空。") from exc

    columns = map_concept_board_columns(headers)
    parsed_rows: list[dict[str, Any]] = []
    for row in rows:
        code = cell_value(row, columns["code"])
        symbol = normalize_a_share_symbol(code)
        name = str(cell_value(row, columns["name"]) or "").strip()
        concepts = str(cell_value(row, columns["concepts"]) or "").strip()
        if not symbol or not name or not concepts:
            continue
        parsed_rows.append(
            {
                "symbol": symbol,
                "name": name,
                "industry": str(cell_value(row, columns.get("industry")) or "").strip(),
                "concepts": concepts,
                "concept_list": split_concepts(concepts),
                "revenue_growth_yoy": to_float(cell_value(row, columns.get("revenue_growth_yoy"))),
                "net_profit_growth_yoy": to_float(cell_value(row, columns.get("net_profit_growth_yoy"))),
                "net_profit_ytd_yi": to_float(cell_value(row, columns.get("net_profit_ytd_yi"))),
                "roe": to_float(cell_value(row, columns.get("roe"))),
                "pe": to_float(cell_value(row, columns.get("pe"))),
                "forecast_pe_2026": to_float(cell_value(row, columns.get("forecast_pe_2026"))),
                "pe_median_since_2023": to_float(cell_value(row, columns.get("pe_median_since_2023"))),
                "forecast_net_profit_growth_yoy": to_float(
                    cell_value(row, columns.get("forecast_net_profit_growth_yoy"))
                ),
                "total_market_cap_cny_100m": to_float(cell_value(row, columns.get("total_market_cap_cny_100m"))),
                "average_trading_price_2025": to_float(cell_value(row, columns.get("average_trading_price_2025"))),
            }
        )
    return tuple(parsed_rows)


XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XLSX_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def read_xlsx_sheet_values(path: Path, sheet_name: str) -> list[tuple[Any, ...]]:
    with ZipFile(path) as archive:
        sheet_path = find_sheet_path(archive, sheet_name)
        shared_strings = read_shared_strings(archive)
        root = ElementTree.fromstring(archive.read(sheet_path))

    rows: list[tuple[Any, ...]] = []
    for row_element in root.findall(f".//{{{XLSX_MAIN_NS}}}sheetData/{{{XLSX_MAIN_NS}}}row"):
        values: list[Any] = []
        for cell in row_element.findall(f"{{{XLSX_MAIN_NS}}}c"):
            index = cell_index(cell, len(values))
            while len(values) <= index:
                values.append(None)
            values[index] = cell_value_from_xml(cell, shared_strings)
        rows.append(tuple(values))
    return rows


def find_sheet_path(archive: ZipFile, sheet_name: str) -> str:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    rel_id = ""
    for sheet in workbook_root.findall(f".//{{{XLSX_MAIN_NS}}}sheet"):
        if sheet.attrib.get("name") == sheet_name:
            rel_id = sheet.attrib.get(f"{{{XLSX_REL_NS}}}id", "")
            break
    if not rel_id:
        raise ValueError(f"本地 A 股概念板块表缺少工作表：{sheet_name}")

    rels_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for rel in rels_root.findall(f"{{{XLSX_PACKAGE_REL_NS}}}Relationship"):
        if rel.attrib.get("Id") != rel_id:
            continue
        target = rel.attrib.get("Target", "")
        if target.startswith("/"):
            return target.lstrip("/")
        return str(Path("xl") / target).replace("\\", "/")
    raise ValueError(f"本地 A 股概念板块表找不到工作表文件：{sheet_name}")


def read_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root.findall(f"{{{XLSX_MAIN_NS}}}si"):
        strings.append("".join(node.text or "" for node in item.findall(f".//{{{XLSX_MAIN_NS}}}t")))
    return strings


def cell_index(cell: Any, fallback: int) -> int:
    reference = str(cell.attrib.get("r") or "")
    match = re.match(r"([A-Z]+)", reference)
    if not match:
        return fallback
    index = 0
    for char in match.group(1):
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def cell_value_from_xml(cell: Any, shared_strings: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(f".//{{{XLSX_MAIN_NS}}}t"))

    value = cell.find(f"{{{XLSX_MAIN_NS}}}v")
    text = value.text if value is not None else ""
    if cell_type == "s":
        try:
            return shared_strings[int(text)]
        except (ValueError, IndexError):
            return ""
    return text


def map_concept_board_columns(headers: tuple[Any, ...]) -> dict[str, int]:
    normalized = [normalize_header(header) for header in headers]
    columns = {
        "code": require_column(normalized, exact=("代码", "证券代码")),
        "name": require_column(normalized, exact=("证券名称", "名称")),
        "concepts": require_column(normalized, exact=("概念板块",)),
        "industry": find_column(normalized, exact=("行业",)),
        "revenue_growth_yoy": find_column(normalized, startswith=("营业收入同比",)),
        "net_profit_growth_yoy": find_column(normalized, startswith=("归母净利润同比",)),
        "net_profit_ytd_yi": find_column(normalized, startswith=("归母净利润",)),
        "roe": find_column(normalized, startswith=("净资产收益率",)),
        "pe": find_column(normalized, exact=("PE",)),
        "forecast_pe_2026": find_column(normalized, startswith=("预测PE",)),
        "pe_median_since_2023": find_column(normalized, startswith=("区间PE中位值",)),
        "forecast_net_profit_growth_yoy": find_column(normalized, startswith=("预测净利润同比",)),
        "total_market_cap_cny_100m": find_column(normalized, startswith=("总市值",)),
        "average_trading_price_2025": find_column(normalized, startswith=("成交均价",)),
    }
    return {key: value for key, value in columns.items() if value is not None}


def require_column(
    headers: list[str],
    *,
    exact: tuple[str, ...] = (),
    startswith: tuple[str, ...] = (),
) -> int:
    column = find_column(headers, exact=exact, startswith=startswith)
    if column is None:
        names = ", ".join(exact or startswith)
        raise ValueError(f"本地 A 股概念板块表缺少必要列：{names}")
    return column


def find_column(
    headers: list[str],
    *,
    exact: tuple[str, ...] = (),
    startswith: tuple[str, ...] = (),
) -> int | None:
    for index, header in enumerate(headers):
        if header in exact or any(header.startswith(prefix) for prefix in startswith):
            return index
    return None


def match_concept_board_rows(
    rows: list[dict[str, Any]],
    requested_sectors: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    merged: dict[str, dict[str, Any]] = {}
    matched_sectors: list[dict[str, Any]] = []

    for requested in requested_sectors:
        requested_key = normalize_match_text(requested)
        if not requested_key:
            continue
        exact_matches = find_matches_for_sector(rows, requested_key, exact=True)
        sector_matches = exact_matches or find_matches_for_sector(rows, requested_key, exact=False)
        matched_concepts = sorted({concept for match in sector_matches for concept in match["matched_concepts"]})
        matched_sectors.append(
            {
                "requested_sector": requested,
                "match_type": "exact" if exact_matches else ("contains" if sector_matches else "none"),
                "matched_concepts": matched_concepts,
                "candidate_count": len(sector_matches),
            }
        )
        for match in sector_matches:
            symbol = match["row"]["symbol"]
            current = merged.setdefault(
                symbol,
                {
                    "row": match["row"],
                    "requested_sectors": [],
                    "matched_concepts": [],
                },
            )
            if requested not in current["requested_sectors"]:
                current["requested_sectors"].append(requested)
            for concept in match["matched_concepts"]:
                if concept not in current["matched_concepts"]:
                    current["matched_concepts"].append(concept)

    return list(merged.values()), matched_sectors


def find_matches_for_sector(
    rows: list[dict[str, Any]],
    requested_key: str,
    *,
    exact: bool,
) -> list[dict[str, Any]]:
    matches = []
    for row in rows:
        matched_concepts = []
        for concept in row.get("concept_list", []):
            concept_key = normalize_match_text(concept)
            if exact:
                is_match = concept_key == requested_key
            else:
                is_match = requested_key in concept_key
            if is_match:
                matched_concepts.append(concept)
        if matched_concepts:
            matches.append({"row": row, "matched_concepts": matched_concepts})
    return matches


def passes_static_filters(row: dict[str, Any], filters: dict[str, Any]) -> bool:
    name = str(row.get("name") or "").upper()
    concepts = [str(item) for item in row.get("concept_list", []) or []]
    if filters.get("exclude_st", True) and ("ST" in name or "ST股" in concepts):
        return False

    exclude_new_days = int(filters.get("exclude_new_days") or 0)
    if exclude_new_days > 0 and any("次新股" in concept or concept == "新股" for concept in concepts):
        return False

    min_market_cap = to_float(filters.get("min_market_cap_yi"))
    market_cap = to_float(row.get("total_market_cap_cny_100m"))
    if min_market_cap is not None and market_cap is not None and market_cap < min_market_cap:
        return False

    max_pe = to_float(filters.get("max_pe"))
    pe = to_float(row.get("pe"))
    if max_pe is not None and pe is not None and pe > 0 and pe > max_pe:
        return False
    return True


def score_static_candidate(row: dict[str, Any]) -> float:
    score = 50.0
    score += capped(to_float(row.get("roe")), lower=-30, upper=30) * 0.8
    score += capped(to_float(row.get("revenue_growth_yoy")), lower=-80, upper=100) * 0.12
    score += capped(to_float(row.get("net_profit_growth_yoy")), lower=-100, upper=120) * 0.14
    score += capped(to_float(row.get("forecast_net_profit_growth_yoy")), lower=-80, upper=100) * 0.10

    market_cap = to_float(row.get("total_market_cap_cny_100m"))
    if market_cap and market_cap > 0:
        score += min(math.log10(market_cap + 1) * 4, 12)

    pe = to_float(row.get("pe"))
    if pe is None:
        score -= 3
    elif pe <= 0:
        score -= 8
    elif pe <= 30:
        score += 10
    elif pe <= 60:
        score += 4
    else:
        score -= min((pe - 60) / 8, 12)
    return round(max(min(score, 100), 0), 2)


def capped(value: float | None, *, lower: float, upper: float) -> float:
    if value is None:
        return 0.0
    return max(min(value, upper), lower)


def build_static_candidate_reason(
    *,
    requested: list[str],
    matched_concepts: list[str],
    score: float,
) -> str:
    requested_text = "、".join(requested)
    concept_text = "、".join(matched_concepts[:4])
    return f"本地概念板块匹配：{requested_text} -> {concept_text}；Excel 静态评分 {score}"


def normalize_a_share_symbol(value: Any) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return ""
    if re.fullmatch(r"\d{6}\.(SH|SZ|BJ)", text):
        return text
    if re.fullmatch(r"(SH|SZ|BJ)\d{6}", text):
        return f"{text[2:]}.{text[:2]}"
    if not re.fullmatch(r"\d{6}", text):
        return text
    if text.startswith(("600", "601", "603", "605", "688")):
        return f"{text}.SH"
    if text.startswith(("4", "8", "9")):
        return f"{text}.BJ"
    return f"{text}.SZ"


def split_concepts(value: str) -> list[str]:
    concepts = []
    for item in re.split(r"[,，]", value):
        concept = item.strip()
        if concept and concept not in concepts:
            concepts.append(concept)
    return concepts


def normalize_match_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).casefold()


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def cell_value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if text in {"", "--", "None", "nan"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None
